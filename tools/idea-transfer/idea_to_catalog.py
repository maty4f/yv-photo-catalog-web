#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
idea_to_catalog.py — transfer catalog data out of an IDEA database into the
yv-photo-catalog web-edition record format.

Run this on the machine that has IDEA (the web session in the cloud cannot
reach a local database). It connects to the source, applies the field mapping
you describe in a JSON config, and writes an array of catalog records that
match tools/idea-transfer/catalog_schema.json.

    python idea_to_catalog.py --mapping mapping.json --out catalog.json
    python idea_to_catalog.py --mapping mapping.json --dry-run --limit 3

The two source modes (set in mapping.json -> "source.type"):

  * "db"   — a live database connection via a SQLAlchemy URL. Covers SQLite,
             SQL Server, MySQL/MariaDB, PostgreSQL, and MS-Access (ODBC).
             Requires: pip install SQLAlchemy (+ the driver for your DB).
  * "file" — a flat export IDEA can produce: .csv / .tsv / .xlsx / .json.
             CSV/TSV/JSON need nothing extra; .xlsx needs: pip install openpyxl.

Nothing here talks to the network — it only reads your source and writes a file.
"""

import argparse
import csv
import json
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
SCHEMA_PATH = os.path.join(HERE, "catalog_schema.json")
THESAURUS_PATH = os.path.normpath(
    os.path.join(HERE, "..", "..", "data", "thesaurus_photo_archive.json")
)

# Fields that carry a house default when the source has no value for them.
# Kept in sync with catalog_schema.json.
SCHEMA_DEFAULTS = {
    "material": "26T קובץ סריקה",
    "rights": "הארכיון",
    "classification": "בלתי מסווג",
    "negative": "ל",
    "size": "רגיל",
    "www": "Y",
    "intranet": "Y",
    "copies": "סרוק",
}

# Every scalar field the record may carry (list/table fields handled separately).
SCALAR_FIELDS = [
    "source_id", "title_he", "title_en",
    "material", "color", "rights", "classification", "negative", "photographer",
    "creation_date", "recon_date", "size", "www", "intranet", "copies",
    "donor_he", "donor_en", "source_he", "source_en",
    "visual_he", "visual_en", "context_he", "context_en",
    "persons_he", "persons_en", "objects_he", "objects_en",
    "inscriptions_he", "inscriptions_en", "studio_he", "studio_en",
    "bio_he", "bio_en", "notes_he", "notes_en",
]


def die(msg):
    sys.stderr.write("error: " + msg + "\n")
    sys.exit(1)


# --------------------------------------------------------------------------- #
#  Reading the source                                                         #
# --------------------------------------------------------------------------- #
def read_db(db_url, query, table):
    """Yield rows (as dicts) from a SQLAlchemy-reachable database."""
    try:
        from sqlalchemy import create_engine, text
    except ImportError:
        die("source.type is 'db' but SQLAlchemy is not installed.\n"
            "       pip install SQLAlchemy  (and the driver for your DB, e.g. pyodbc)")
    if not db_url or str(db_url).startswith("REPLACE"):
        die("source.db_url is not set — edit mapping.json and put the real "
            "SQLAlchemy connection URL for your IDEA database.")
    if not query:
        if not table:
            die("source needs either a 'query' or a 'table'.")
        query = "SELECT * FROM " + table
    engine = create_engine(db_url)
    with engine.connect() as conn:
        result = conn.execute(text(query))
        keys = list(result.keys())
        for row in result:
            yield dict(zip(keys, row))


def read_file(path, sheet, encoding):
    """Yield rows (as dicts) from a csv/tsv/xlsx/json export."""
    if not path or not os.path.exists(path):
        die("source.file not found: %r (path is relative to where you run the tool)" % path)
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv", ".tsv", ".txt"):
        delim = "\t" if ext == ".tsv" else ","
        with open(path, "r", encoding=encoding, newline="") as fh:
            for row in csv.DictReader(fh, delimiter=delim):
                yield row
    elif ext == ".json":
        with open(path, "r", encoding=encoding) as fh:
            data = json.load(fh)
        rows = data if isinstance(data, list) else data.get("records") or data.get("data") or []
        for row in rows:
            yield row
    elif ext == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:
            die("reading .xlsx needs openpyxl:  pip install openpyxl")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header = [str(c) if c is not None else "" for c in next(rows)]
        except StopIteration:
            return
        for r in rows:
            yield {header[i]: r[i] for i in range(min(len(header), len(r)))}
    else:
        die("unsupported file type: %s (use .csv/.tsv/.xlsx/.json)" % ext)


def read_source(src):
    typ = (src.get("type") or "db").lower()
    if typ == "db":
        return read_db(src.get("db_url"), src.get("query"), src.get("table"))
    if typ == "file":
        return read_file(src.get("file"), src.get("sheet"), src.get("encoding") or "utf-8-sig")
    die("source.type must be 'db' or 'file', got %r" % typ)


# --------------------------------------------------------------------------- #
#  Mapping a source row -> a catalog record                                   #
# --------------------------------------------------------------------------- #
def cell(row, column):
    """Case-insensitive column lookup; returns '' for missing/None."""
    if column in row:
        v = row[column]
    else:
        low = {str(k).lower(): k for k in row}
        key = low.get(str(column).lower())
        v = row[key] if key is not None else None
    if v is None:
        return ""
    return str(v).strip()


def apply_field(row, spec):
    """Resolve one field spec ({column|const|template}, optional split) to a value."""
    if "const" in spec:
        value = spec["const"]
    elif "template" in spec:
        value = spec["template"]
        for k in list(row.keys()):
            value = value.replace("{%s}" % k, cell(row, k))
    elif "column" in spec:
        value = cell(row, spec["column"])
    else:
        return ""
    if isinstance(value, str) and spec.get("strip", True):
        value = value.strip()
    if "split" in spec and isinstance(value, str):
        sep = spec["split"]
        return [p.strip() for p in value.split(sep) if p.strip()]
    return value


def load_thesaurus():
    """name(HE, lowercased) -> {'he':.., 'en':..} for subject matching."""
    idx = {}
    try:
        with open(THESAURUS_PATH, "r", encoding="utf-8") as fh:
            for e in json.load(fh):
                if e.get("he"):
                    idx[e["he"].strip().lower()] = {"he": e["he"], "en": e.get("en", "")}
                if e.get("en"):
                    idx.setdefault(e["en"].strip().lower(), {"he": e.get("he", ""), "en": e["en"]})
    except (OSError, ValueError):
        pass
    return idx


def build_subjects(row, cfg, thes):
    if not cfg or "column" not in cfg:
        return []
    terms = apply_field(row, {"column": cfg["column"], "split": cfg.get("split", ";")})
    if isinstance(terms, str):
        terms = [terms] if terms else []
    out = []
    match = cfg.get("thesaurus") == "match"
    for t in terms:
        hit = thes.get(t.strip().lower()) if match else None
        out.append(dict(hit) if hit else {"he": t, "en": ""})
    return out


def build_people(row, cfg):
    if not cfg or cfg.get("mode", "none") == "none":
        return []
    if cfg["mode"] == "inline":
        ic = cfg.get("inline", {})
        names = apply_field(row, {"column": ic.get("column"), "split": ic.get("split", ";")})
        if isinstance(names, str):
            names = [names] if names else []
        return [{"name": n, "position_he": "", "position_en": ""} for n in names]
    # mode == "join" is resolved in main() where the join table is pre-loaded.
    return []


def build_record(row, mapping, thes, defaults):
    rec = {}
    fields = mapping.get("fields", {})
    for name in SCALAR_FIELDS:
        if name in fields:
            rec[name] = apply_field(row, fields[name])
        elif name in defaults:
            rec[name] = defaults[name]
    # places (list field with a default of [])
    if "places" in fields:
        pv = apply_field(row, fields["places"])
        rec["places"] = pv if isinstance(pv, list) else ([pv] if pv else [])
    rec["subjects"] = build_subjects(row, mapping.get("subjects"), thes)
    rec["people"] = build_people(row, mapping.get("people"))
    # drop empty scalars so records stay lean, but always keep title fields
    return {k: v for k, v in rec.items()
            if v not in ("", [], None) or k in ("title_he", "title_en")}


# --------------------------------------------------------------------------- #
#  People join table (optional)                                               #
# --------------------------------------------------------------------------- #
def load_people_join(pcfg):
    """Return {record_id -> [ {name, position_he, position_en}, ... ]} or None."""
    if not pcfg or pcfg.get("mode") != "join":
        return None
    j = pcfg.get("join", {})
    rows = read_db(j.get("db_url"), j.get("query"), j.get("table"))
    by_id = {}
    for r in rows:
        rid = cell(r, j.get("key", "record_id"))
        by_id.setdefault(rid, []).append({
            "name": cell(r, j.get("name", "full_name")),
            "position_he": cell(r, j.get("position_he", "position_he")),
            "position_en": cell(r, j.get("position_en", "position_en")),
        })
    return by_id


# --------------------------------------------------------------------------- #
#  Main                                                                        #
# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(
        description="Transfer catalog data from an IDEA database into the "
                    "yv-photo-catalog record format.")
    ap.add_argument("--mapping", required=True, help="path to the mapping JSON (copy mapping.example.json)")
    ap.add_argument("--out", help="output JSON file (default: stdout)")
    ap.add_argument("--ndjson", action="store_true", help="write newline-delimited JSON instead of a JSON array")
    ap.add_argument("--limit", type=int, default=0, help="stop after N source rows (0 = all)")
    ap.add_argument("--dry-run", action="store_true", help="print mapped records to stderr, write nothing")
    args = ap.parse_args()

    with open(args.mapping, "r", encoding="utf-8") as fh:
        mapping = json.load(fh)

    defaults = dict(SCHEMA_DEFAULTS)
    defaults.update(mapping.get("defaults", {}))
    thes = load_thesaurus()
    people_join = load_people_join(mapping.get("people"))
    pkey_field = (mapping.get("people", {}).get("join", {}).get("record_id_field")
                  or mapping.get("fields", {}).get("source_id", {}).get("column"))

    records = []
    n = 0
    for row in read_source(mapping.get("source", {})):
        rec = build_record(row, mapping, thes, defaults)
        if people_join is not None:
            rid = cell(row, pkey_field) if pkey_field else rec.get("source_id", "")
            if rid in people_join:
                rec["people"] = people_join[rid]
        records.append(rec)
        n += 1
        if args.dry_run and n <= (args.limit or 3):
            sys.stderr.write(json.dumps(rec, ensure_ascii=False, indent=2) + "\n")
        if args.limit and n >= args.limit:
            break

    sys.stderr.write("mapped %d record(s)\n" % len(records))
    with_title = sum(1 for r in records if r.get("title_he") or r.get("title_en"))
    if with_title < len(records):
        sys.stderr.write("warning: %d record(s) have no title — check your title_he/title_en mapping\n"
                         % (len(records) - with_title))
    if args.dry_run:
        sys.stderr.write("dry-run: nothing written\n")
        return

    def dump(fh):
        if args.ndjson:
            for r in records:
                fh.write(json.dumps(r, ensure_ascii=False) + "\n")
        else:
            json.dump(records, fh, ensure_ascii=False, indent=2)
            fh.write("\n")

    if args.out:
        with open(args.out, "w", encoding="utf-8") as fh:
            dump(fh)
        sys.stderr.write("wrote %s\n" % args.out)
    else:
        dump(sys.stdout)


if __name__ == "__main__":
    main()
